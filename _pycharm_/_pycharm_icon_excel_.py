import os
import time
import string

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from svglib.svglib import svg2rlg
from reportlab.graphics import renderPM

column = list(string.ascii_uppercase)
column += [f"{a}{b}" for a in column for b in column]
row = [f"{index}" for index in range(1, len(column))]

icon_path = os.path.join(os.getcwd(), "icons", "toolwindows")
excel_path = os.path.join(os.getcwd(), "test.xlsx")


def set_column_row():
    wb = load_workbook(excel_path)
    ws = wb.active
    for item in column:
        ws.column_dimensions[item].width = 4
    for item in row:
        ws.row_dimensions[item].height = 4 * 5
    wb.save(excel_path)
    wb.close()


def main():
    wb = load_workbook(excel_path)
    ws = wb.active
    xy = [f"A{index}" for index in range(1, 999)]
    for image_index, image_file in zip(xy, os.listdir(icon_path)):
        image_path = os.path.join(icon_path, image_file)
        if os.path.isdir(image_path):
            continue
        if image_path.endswith(".svg"):
            path_old = image_path
            path_new = image_path.replace(".svg", ".png")
            drawing = svg2rlg(path_old)
            renderPM.drawToFile(drawing, path_new, fmt='PNG')
            image_path = path_new
            print(image_index, image_path)
        img = Image(image_path)
        img.width, img.height = 25, 25
        ws.add_image(img, image_index)
    wb.save(excel_path)
    wb.close()


if __name__ == '__main__':
    set_column_row()
    time.sleep(5)
    main()
